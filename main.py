# Variabeln
## Pfad zum gewünschten Excel Dokument
path_to_file = "Liste KNX def 16_3_22.xlsx"

from openpyxl import load_workbook

# Hauptfuntion die (momentan) das Dokument in eine Python Datenstruktur bringt
def main():
    file, sheet, = import_file(path_to_file)
    data = get_data(sheet)
    display_all(data)



# Testfunktion die alle Geschosse mit allen Aktionen und Messwerten anzeigt
def display_all(data):
    for geschoss in data:
        print()
        print("----------------------------------------------")
        print("Geschoss: {}".format(geschoss["name"]))
        for aktion in geschoss["aktion"]:
            print("----------------------------------------------")
            print("Aktion: {}".format(aktion["name"]))
            for messwert in aktion["messwert"]:
                print(messwert["name"])
        print("----------------------------------------------")

# Importiert das Dokument und gibt dieses sowie das aktuelle Blatt zurück
def import_file(path):
    wb = load_workbook(filename = path) 
    ws = wb.active
    return wb, ws 

# Wandelt das Blatt in eine Python Datenstruktur bringt
def get_data(sheet):
    i_geschoss, i_aktion = 1, 0
    data = []
    for row in sheet.iter_rows(min_col = 5, max_col = 5, max_row = 2000):
        for cell in row:
            if cell.value == None:
                continue
            split = str(cell.value).split("/")
            if split[0] == str(i_geschoss) and split[1] == "-":
                name = sheet.cell(row[0].row, column = 1).value
                geschoss = {"name": name, 
                            "id_geschoss": i_geschoss, 
                            "aktion": []}
                data.append(geschoss)
                i_geschoss += 1
                i_aktion = 0
            elif split[1] >= str(i_aktion) and split[2] == "-":
                i_aktion = int(split[1])
                name = sheet.cell(row[0].row, column = 2).value
                if sheet.cell(row[0].row, column=6).value == "true":
                    bool_f = True 
                else:
                    bool_f = False
                aktion = {"name": name, 
                          "id_aktion": i_aktion, 
                          "bool_f": bool_f,
                          "messwert": []}
                data[-1]["aktion"].append(aktion)
                i_aktion += 1
            else:
                name = sheet.cell(row[0].row, column = 3).value
                if sheet.cell(row[0].row, column=6).value == "true":
                    aufschalten = True 
                else:
                    aufschalten = False
                if sheet.cell(row[0].row, column=6).value == "true":
                    bool_f = True 
                else:
                    bool_f = False
                kommentar = sheet.cell(row[0].row, column = 7).value
                dpst = sheet.cell(row[0].row, column = 8).value
                messwert = {"name": name,
                            "aufschalten": aufschalten,
                            "id_messwert": int(split[2]),
                            "bool_f": bool_f,
                            "kommentar": kommentar,
                            "dpst": dpst, }               
                #data[i_geschoss-2]["aktion"]["messwert"]
                data[-1]["aktion"][-1]["messwert"].append(messwert)
    return data

if __name__ == '__main__':
  main()
